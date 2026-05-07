"""Load a QuickBooks transactions Excel export into the same record schema as the parsers.

The QB format we expect (validated against Allegiance.xlsx):
- Row 1: column header (Type, Date, Num, Memo, Due Date, Open Balance, Amount)
- Row 2: account info (e.g. "ALLEGIANCE TRUCKS ACCT 40981")
- Rows 3..n-2: data rows
- Last 2 rows: totals (skip)

Columns are at fixed letters: E=Type, G=Date, I=Num, K=Memo, M=Due Date, O=OpenBal, Q=Amount
(QB exports typically have empty padding columns between data columns)
"""
from __future__ import annotations
import datetime
from typing import List, Optional
from openpyxl import load_workbook

from .parsers.base import StatementRecord, ParsedStatement


def load_qb(xlsx_path: str, sheet: Optional[str] = None) -> ParsedStatement:
    wb = load_workbook(xlsx_path, data_only=True)
    sh = wb[sheet] if sheet else wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    # Detect column letters from header row
    header = list(sh.iter_rows(min_row=1, max_row=1, values_only=False))[0]
    col_map = {}
    for cell in header:
        v = cell.value
        if not v:
            continue
        v = str(v).strip().lower()
        if v == "type": col_map["type"] = cell.column
        elif v == "date": col_map["date"] = cell.column
        elif v == "num": col_map["num"] = cell.column
        elif v == "memo": col_map["memo"] = cell.column
        elif v == "due date": col_map["due"] = cell.column
        elif v == "open balance": col_map["open"] = cell.column
        elif v == "amount": col_map["amt"] = cell.column

    # Account name from row 2 column B (best effort)
    account = None
    if sh.max_row >= 2:
        for cell in sh[2]:
            if cell.value and isinstance(cell.value, str) and cell.value.strip():
                account = cell.value.strip()
                break

    records: List[StatementRecord] = []
    # Iterate all rows from 3 onward; skip rows where Type column is empty (header pads / totals)
    for row in sh.iter_rows(min_row=3, values_only=False):
        get = lambda key: row[col_map[key] - 1].value if key in col_map else None
        t = get("type")
        if not t:
            continue
        num = get("num")
        if not num:
            continue
        d = get("date")
        amt = get("amt")
        memo = get("memo")
        if isinstance(d, datetime.datetime):
            d = d.date()
        try:
            amt_v = float(amt) if amt is not None else 0.0
        except (TypeError, ValueError):
            amt_v = 0.0
        records.append(StatementRecord(
            invoice_no=str(num).strip(),
            txn_date=d,
            amount=amt_v,
            type=str(t).strip(),
            po_ref=None,
            raw=str(memo or "").strip(),
        ))

    return ParsedStatement(
        vendor="QuickBooks",
        statement_date=None,
        account_no=account,
        period_total=round(sum(r.amount for r in records), 2),
        statement_mode="all_activity",
        records=records,
        source_file=xlsx_path,
    )


def merge_qb_statements(stmts: list[ParsedStatement]) -> ParsedStatement:
    """Combine multiple QB exports into one ParsedStatement for reconciliation."""
    if len(stmts) == 1:
        return stmts[0]
    all_records = [r for s in stmts for r in s.records]
    accounts = ", ".join(filter(None, (s.account_no for s in stmts)))
    sources  = ", ".join(filter(None, (s.source_file for s in stmts)))
    return ParsedStatement(
        vendor="QuickBooks",
        statement_date=None,
        account_no=accounts or None,
        period_total=round(sum(r.amount for r in all_records), 2),
        statement_mode="all_activity",
        records=all_records,
        source_file=sources or None,
    )


if __name__ == "__main__":
    import sys
    qb = load_qb(sys.argv[1])
    print(f"Account: {qb.account_no}")
    print(f"Records: {len(qb.records)}")
    print(f"Sum: {qb.computed_total()}")
    for r in qb.records[:5]:
        print(f"  {r.txn_date} | {r.invoice_no} | {r.amount:>10.2f} | {r.type}")
