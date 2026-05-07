"""Compare a vendor ParsedStatement against a QuickBooks ParsedStatement.

Matching strategy: normalize invoice numbers (strip leading zeros, lowercase),
then join on that key. One-to-one on the first occurrence; duplicates (e.g.
partial payments) fall into the unmatched buckets.

Discrepancy logic:
  - amount mismatch  : abs(stmt_amount - qb_amount) >= 0.02
  - date flag        : abs(stmt_date - qb_date) > 1 day (tolerance for posting lag)
  Clean match requires both conditions clear.
"""
from __future__ import annotations

import io
import json
from dataclasses import dataclass, field
from datetime import date
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .parsers.base import ParsedStatement, StatementRecord


@dataclass
class MatchedRecord:
    invoice_no: str
    stmt_date: Optional[date]
    qb_date: Optional[date]
    stmt_amount: float
    qb_amount: float
    delta: float                        # stmt_amount - qb_amount
    date_drift_days: int                # 0 if either date is None
    discrepancy_reason: Optional[str]   # None | "amount" | "date" | "amount+date"


@dataclass
class FuzzyMatchedRecord:
    """A pair matched by amount agreement + near-identical invoice number (edit distance = 1).
    Requires human review — likely a typo in one system."""
    invoice_no_stmt: str    # as it appears on the vendor statement
    invoice_no_qb: str      # as it appears in QuickBooks
    stmt_date: Optional[date]
    qb_date: Optional[date]
    stmt_amount: float
    qb_amount: float
    delta: float
    date_drift_days: int
    edit_distance: int      # Levenshtein distance between normalized invoice numbers


@dataclass
class ReconciliationResult:
    vendor: str
    statement_mode: str                 # "open_only" | "all_activity"
    stmt_total: Optional[float]
    qb_total: float
    stmt_record_count: int
    qb_record_count: int
    matched: List[MatchedRecord] = field(default_factory=list)
    discrepancies: List[MatchedRecord] = field(default_factory=list)
    fuzzy_matches: List[FuzzyMatchedRecord] = field(default_factory=list)
    stmt_only: List[StatementRecord] = field(default_factory=list)
    qb_only: List[StatementRecord] = field(default_factory=list)

    @property
    def match_rate(self) -> float:
        denom = len(self.matched) + len(self.discrepancies) + len(self.stmt_only)
        return round(len(self.matched) / denom, 4) if denom else 0.0

    def to_json(self) -> bytes:
        def _d(v) -> Optional[str]:
            return v.isoformat() if isinstance(v, date) else None

        data = {
            "vendor": self.vendor,
            "statement_mode": self.statement_mode,
            "stmt_total": self.stmt_total,
            "qb_total": self.qb_total,
            "stmt_record_count": self.stmt_record_count,
            "qb_record_count": self.qb_record_count,
            "matched": [
                {
                    "invoice_no": r.invoice_no,
                    "stmt_date": _d(r.stmt_date),
                    "qb_date": _d(r.qb_date),
                    "stmt_amount": r.stmt_amount,
                    "qb_amount": r.qb_amount,
                    "delta": r.delta,
                    "date_drift_days": r.date_drift_days,
                    "reason": r.discrepancy_reason,
                }
                for r in self.matched
            ],
            "discrepancies": [
                {
                    "invoice_no": r.invoice_no,
                    "stmt_date": _d(r.stmt_date),
                    "qb_date": _d(r.qb_date),
                    "stmt_amount": r.stmt_amount,
                    "qb_amount": r.qb_amount,
                    "delta": r.delta,
                    "date_drift_days": r.date_drift_days,
                    "reason": r.discrepancy_reason,
                }
                for r in self.discrepancies
            ],
            "stmt_only": [
                {
                    "invoice_no": r.invoice_no,
                    "date": _d(r.txn_date),
                    "amount": r.amount,
                    "type": r.type,
                    "po_ref": r.po_ref,
                    "source": "statement",
                }
                for r in self.stmt_only
            ],
            "qb_only": [
                {
                    "invoice_no": r.invoice_no,
                    "date": _d(r.txn_date),
                    "amount": r.amount,
                    "type": r.type,
                    "source": "qb",
                }
                for r in self.qb_only
            ],
            "fuzzy_matches": [
                {
                    "invoice_no_stmt": r.invoice_no_stmt,
                    "invoice_no_qb":   r.invoice_no_qb,
                    "stmt_date":       _d(r.stmt_date),
                    "qb_date":         _d(r.qb_date),
                    "stmt_amount":     r.stmt_amount,
                    "qb_amount":       r.qb_amount,
                    "delta":           r.delta,
                    "date_drift_days": r.date_drift_days,
                    "edit_distance":   r.edit_distance,
                }
                for r in self.fuzzy_matches
            ],
        }
        return json.dumps(data, indent=2).encode()


def reconcile(stmt: ParsedStatement, qb: ParsedStatement) -> ReconciliationResult:
    """Match stmt records against QB records and bucket them by outcome."""
    stmt_map: dict[str, list[StatementRecord]] = {}
    for r in stmt.records:
        stmt_map.setdefault(_norm(r.invoice_no), []).append(r)

    qb_map: dict[str, list[StatementRecord]] = {}
    for r in qb.records:
        qb_map.setdefault(_norm(r.invoice_no), []).append(r)

    matched: List[MatchedRecord] = []
    discrepancies: List[MatchedRecord] = []
    matched_keys: set[str] = set()

    for key, s_list in stmt_map.items():
        if key not in qb_map:
            continue
        matched_keys.add(key)
        s = s_list[0]
        q = qb_map[key][0]

        delta = round(s.amount - q.amount, 2)
        drift = _date_drift(s.txn_date, q.txn_date)

        amount_ok = abs(delta) < 0.02
        date_ok   = drift <= 1

        if amount_ok and date_ok:
            reason = None
        elif not amount_ok and not date_ok:
            reason = "amount+date"
        elif not amount_ok:
            reason = "amount"
        else:
            reason = "date"

        mr = MatchedRecord(
            invoice_no=s.invoice_no,
            stmt_date=s.txn_date,
            qb_date=q.txn_date,
            stmt_amount=s.amount,
            qb_amount=q.amount,
            delta=delta,
            date_drift_days=drift,
            discrepancy_reason=reason,
        )
        (matched if reason is None else discrepancies).append(mr)

    # ── Fuzzy matching pass ──────────────────────────────────────────────────
    # For records that didn't match on invoice number, try matching by
    # amount + near-identical invoice number (edit distance = 1 = likely typo).
    fuzzy_matches: List[FuzzyMatchedRecord] = []
    fuzzy_stmt_keys: set[str] = set()
    fuzzy_qb_keys:   set[str] = set()

    unmatched_stmt = [(k, rs[0]) for k, rs in stmt_map.items() if k not in matched_keys]
    unmatched_qb   = {k: rs[0]  for k, rs in qb_map.items()   if k not in matched_keys}

    for stmt_key, s in unmatched_stmt:
        candidates = [
            (qk, q)
            for qk, q in unmatched_qb.items()
            if qk not in fuzzy_qb_keys
            and abs(s.amount - q.amount) < 0.02
            and _levenshtein(stmt_key, qk) == 1
        ]
        if len(candidates) != 1:   # ambiguous or no near-match
            continue
        qb_key, q = candidates[0]
        fuzzy_stmt_keys.add(stmt_key)
        fuzzy_qb_keys.add(qb_key)
        delta = round(s.amount - q.amount, 2)
        drift = _date_drift(s.txn_date, q.txn_date)
        fuzzy_matches.append(FuzzyMatchedRecord(
            invoice_no_stmt=s.invoice_no,
            invoice_no_qb=q.invoice_no,
            stmt_date=s.txn_date,
            qb_date=q.txn_date,
            stmt_amount=s.amount,
            qb_amount=q.amount,
            delta=delta,
            date_drift_days=drift,
            edit_distance=1,
        ))

    stmt_only = [r for k, rs in stmt_map.items() if k not in matched_keys and k not in fuzzy_stmt_keys for r in rs]
    qb_only   = [r for k, rs in qb_map.items()  if k not in matched_keys and k not in fuzzy_qb_keys   for r in rs]

    return ReconciliationResult(
        vendor=stmt.vendor,
        statement_mode=getattr(stmt, "statement_mode", "open_only"),
        stmt_total=stmt.period_total,
        qb_total=round(sum(r.amount for r in qb.records), 2),
        stmt_record_count=len(stmt.records),
        qb_record_count=len(qb.records),
        matched=matched,
        discrepancies=discrepancies,
        fuzzy_matches=fuzzy_matches,
        stmt_only=stmt_only,
        qb_only=qb_only,
    )


def _levenshtein(a: str, b: str) -> int:
    """Standard dynamic-programming Levenshtein distance."""
    if a == b:
        return 0
    if len(a) < len(b):
        a, b = b, a
    row = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        new_row = [i]
        for j, cb in enumerate(b, 1):
            new_row.append(min(row[j] + 1, new_row[j - 1] + 1, row[j - 1] + (ca != cb)))
        row = new_row
    return row[-1]


def _norm(inv: str) -> str:
    # Remove dashes before stripping leading zeros so formats like
    # "0000-2973846" and "2973846" resolve to the same key.
    return str(inv).strip().replace("-", "").lstrip("0").lower()


def _date_drift(d1, d2) -> int:
    if d1 is None or d2 is None:
        return 0
    try:
        return abs((d1 - d2).days)
    except Exception:
        return 0


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

_HDR_FILL     = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT     = Font(bold=True, color="FFFFFF")
_DISC_FILL    = PatternFill("solid", fgColor="FFE0E0")   # red tint — discrepancies
_DRIFT_FILL   = PatternFill("solid", fgColor="FFF3CD")   # amber tint — date drift warning
_MISSING_FILL = PatternFill("solid", fgColor="FDEBD0")   # orange tint — unmatched
_FUZZY_FILL   = PatternFill("solid", fgColor="EDE7F6")   # lavender — likely matches (needs review)


def _header(ws, cols: list[str]) -> None:
    ws.append(cols)
    for cell in ws[ws.max_row]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def export_to_excel(result: ReconciliationResult) -> bytes:
    wb = openpyxl.Workbook()

    # ── Summary ──────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _set_col_widths(ws, [32, 18])

    activity_note = (
        "all_activity — stated total is current balance, not period sum"
        if result.statement_mode == "all_activity" else ""
    )
    period_diff = round((result.stmt_total or 0) - result.qb_total, 2) if result.stmt_total else None

    summary_rows = [
        ("Vendor",                          result.vendor),
        ("Statement Mode",                  result.statement_mode + (f" ({activity_note})" if activity_note else "")),
        ("",                                ""),
        ("── Period Total Comparison ──",   ""),
        ("Statement Total (stated)",        result.stmt_total),
        ("QB Records Total",                result.qb_total),
        ("Difference (Stmt − QB)",          period_diff),
        ("",                                ""),
        ("── Record Counts ──",             ""),
        ("Statement records parsed",        result.stmt_record_count),
        ("QB records loaded",               result.qb_record_count),
        ("",                                ""),
        ("── Match Results ──",             ""),
        ("Clean Matches",                   len(result.matched)),
        ("Discrepancies",                   len(result.discrepancies)),
        ("Likely Matches (needs review)",   len(result.fuzzy_matches)),
        ("Statement Only (not in QB)",      len(result.stmt_only)),
        ("QB Only (not on statement)",      len(result.qb_only)),
        ("Match Rate",                      f"{result.match_rate:.1%}"),
    ]
    for r in summary_rows:
        ws.append(list(r))

    # Bold the section headers
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("──"):
                cell.font = Font(bold=True)

    # ── Discrepancies ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Discrepancies")
    _header(ws2, ["Invoice No", "Stmt Date", "QB Date", "Date Drift (days)",
                  "Stmt Amount", "QB Amount", "Delta", "Reason"])
    _set_col_widths(ws2, [16, 12, 12, 18, 14, 14, 12, 14])
    for r in result.discrepancies:
        row_idx = ws2.max_row + 1
        ws2.append([
            r.invoice_no,
            str(r.stmt_date or ""),
            str(r.qb_date or ""),
            r.date_drift_days,
            r.stmt_amount,
            r.qb_amount,
            r.delta,
            r.discrepancy_reason or "",
        ])
        for cell in ws2[row_idx]:
            cell.fill = _DISC_FILL

    # ── Matched ──────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Matched")
    _header(ws3, ["Invoice No", "Stmt Date", "QB Date", "Date Drift (days)", "Amount"])
    _set_col_widths(ws3, [16, 12, 12, 18, 14])
    for r in result.matched:
        row_idx = ws3.max_row + 1
        ws3.append([
            r.invoice_no,
            str(r.stmt_date or ""),
            str(r.qb_date or ""),
            r.date_drift_days,
            r.stmt_amount,
        ])
        # Amber highlight if dates disagree (within tolerance but non-zero)
        if r.date_drift_days > 0:
            for cell in ws3[row_idx]:
                cell.fill = _DRIFT_FILL

    # ── Likely Matches (fuzzy — needs human review) ──────────────────────────
    ws4 = wb.create_sheet("Likely Matches")
    _header(ws4, ["Stmt Invoice No", "QB Invoice No", "Stmt Date", "QB Date",
                  "Date Drift (days)", "Amount", "Delta", "Edit Distance"])
    _set_col_widths(ws4, [18, 18, 12, 12, 18, 14, 12, 14])
    for r in result.fuzzy_matches:
        row_idx = ws4.max_row + 1
        ws4.append([
            r.invoice_no_stmt,
            r.invoice_no_qb,
            str(r.stmt_date or ""),
            str(r.qb_date or ""),
            r.date_drift_days,
            r.stmt_amount,
            r.delta,
            r.edit_distance,
        ])
        for cell in ws4[row_idx]:
            cell.fill = _FUZZY_FILL

    # ── Unmatched (stmt_only + qb_only combined) ─────────────────────────────
    ws5 = wb.create_sheet("Unmatched")
    _header(ws5, ["Invoice No", "Date", "Amount", "Type", "PO Ref", "Source"])
    _set_col_widths(ws5, [16, 12, 14, 10, 16, 12])
    for r in result.stmt_only:
        row_idx = ws5.max_row + 1
        ws5.append([r.invoice_no, str(r.txn_date or ""), r.amount, r.type, r.po_ref or "", "Statement"])
        for cell in ws5[row_idx]:
            cell.fill = _MISSING_FILL
    for r in result.qb_only:
        row_idx = ws5.max_row + 1
        ws5.append([r.invoice_no, str(r.txn_date or ""), r.amount, r.type, "", "QB"])
        for cell in ws5[row_idx]:
            cell.fill = _MISSING_FILL

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
